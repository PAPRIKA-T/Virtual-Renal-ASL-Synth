import math
from enum import Enum
from PyQt6.QtGui import QImage,QPixmap
import os
import re
import openpyxl
import argparse


parser = argparse.ArgumentParser()
parser.add_argument('--PredInterval', type=int, default=1000)
parser.add_argument('--PredNiiDir', type=str, default=r'./Resource/DATA/pred_nii_best/')
parser.add_argument('--PredExcel', type=str, default=r'./Resource/pred_nii_best_metric.xlsx')
parser.add_argument('--PredNiiSaveDir', type=str, default=r'./Resource/DATA/')
parser.add_argument('--WindowICON', type=str, default=r'./Resource/ICON.png')
VRAS_config = parser.parse_args()

class AnatomicalPlane(Enum):
    AXIAL = 0       # 水平面
    CORONAL = 1     # 冠状面
    SAGITTAL = 2    # 矢状面

def array_to_qimage(image_array):
    if len(image_array.shape) == 2:  # 灰度图像
        height, width = image_array.shape
        qimage = QImage(image_array.data, width, height, width, QImage.Format.Format_Grayscale8)
    elif len(image_array.shape) == 3 and image_array.shape[2] == 3:  # RGB 彩色图像
        height, width, channels = image_array.shape
        qimage = QImage(image_array.tobytes(), width, height, width, QImage.Format.Format_RGB888)
    else:
        raise ValueError("Unsupported image format")
    return qimage


def array_to_qpixmap(image_array):
    qimage = array_to_qimage(image_array)
    qpixmap = QPixmap.fromImage(qimage)
    return qpixmap

def search_files(directory, search_string):
    matches = []
    # Compile the search string into a regex pattern
    pattern = re.compile(search_string)

    # Traverse the directory
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            # print('file_path:', file_path)
            if search_string in file:
                file_path = os.path.join(root, file)
                matches.append(file_path)
            else:
                print('search_files(): pass')
            pass
    return matches

def search_excel_column(file_path, column, search_string):
    matches = []
    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    # Iterate over the specified column
    for row in range(1, sheet.max_row + 1):
        cell = sheet[f"{column}{row}"]
        if search_string in str(cell.value):
            matches.append((row, cell.value))
    return matches


def read_excel_row_values(file_path, row_number):
    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Retrieve all column values in the specified row
    row_values = [sheet.cell(row=row_number, column=col).value for col in range(1, sheet.max_column + 1)]

    return row_values


def get_magnitude(num):
    if num == 0:
        return float('-inf')  # 0 的数量级定义为负无穷大

    magnitude = math.floor(math.log10(abs(num)))
    return magnitude
